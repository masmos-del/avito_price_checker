import logging
import os
from datetime import datetime
from io import BytesIO
import csv
import json

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

from models import get_session, Product

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DataExporter:
    """Экспорт данных товаров в разные форматы"""
    
    def __init__(self):
        self.session = get_session()
    
    def get_products_data(self, filters=None):
        """Получить данные товаров с фильтрами"""
        query = self.session.query(Product).order_by(Product.created_at.desc())
        
        if filters:
            if filters.get('min_discount'):
                query = query.filter(Product.discount_percent >= filters['min_discount'])
            
            if filters.get('max_price'):
                query = query.filter(Product.price <= filters['max_price'])
            
            if filters.get('min_rating'):
                query = query.filter(Product.seller_rating >= filters['min_rating'])
            
            if filters.get('profitable_only'):
                query = query.filter(Product.is_profitable == True)
            
            if filters.get('sent_only'):
                query = query.filter(Product.is_sent_to_telegram == True)
        
        products = query.all()
        
        # Преобразуем в список дичионариев
        data = []
        for product in products:
            data.append({
                'ID': product.id,
                'Авито ID': product.avito_id,
                'Название': product.title,
                'Категория': product.category,
                'Цена': product.price,
                'Рыночная цена': product.market_price,
                'Скидка %': product.discount_percent,
                'Выгодно': 'Да' if product.is_profitable else 'Нет',
                'Продавец': product.seller_name,
                'Рейтинг': product.seller_rating,
                'Локация': product.location,
                'Отправлено': 'Да' if product.is_sent_to_telegram else 'Нет',
                'Ссылка': product.url,
                'Дата': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '',
            })
        
        return data
    
    def export_to_csv(self, filters=None, filename=None):
        """Экспортировать в CSV"""
        try:
            logger.info("📄 Экспорт в CSV...")
            
            data = self.get_products_data(filters)
            
            if not data:
                logger.warning("❌ Нет данных для экспорта")
                return None
            
            if not filename:
                filename = f"avito_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            # Пишем CSV
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                if data:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            
            logger.info(f"✅ CSV экспортирован: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта CSV: {e}")
            return None
    
    def export_to_excel(self, filters=None, filename=None):
        """Экспортировать в Excel"""
        try:
            if not HAS_EXCEL:
                logger.error("❌ openpyxl не установлен")
                return None
            
            logger.info("📊 Экспорт в Excel...")
            
            data = self.get_products_data(filters)
            
            if not data:
                logger.warning("❌ Нет данных для экспорта")
                return None
            
            if not filename:
                filename = f"avito_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            # Создаём DataFrame
            df = pd.DataFrame(data)
            
            # Пишем в Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Товары', index=False)
                
                # Стилизация
                workbook = writer.book
                worksheet = writer.sheets['Товары']
                
                # Заголовки
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Ширина колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Границы
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            logger.info(f"✅ Excel экспортирован: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта Excel: {e}")
            return None
    
    def export_to_json(self, filters=None, filename=None):
        """Экспортировать в JSON"""
        try:
            logger.info("📋 Экспорт в JSON...")
            
            data = self.get_products_data(filters)
            
            if not data:
                logger.warning("❌ Нет данных для экспорта")
                return None
            
            if not filename:
                filename = f"avito_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ JSON экспортирован: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта JSON: {e}")
            return None
    
    def export_to_memory(self, format='csv', filters=None):
        """Экспортировать в память (для скачивания)"""
        try:
            data = self.get_products_data(filters)
            
            if not data:
                return None
            
            if format == 'csv':
                output = BytesIO()
                df = pd.DataFrame(data)
                df.to_csv(output, index=False, encoding='utf-8-sig')
                output.seek(0)
                return output
            
            elif format == 'excel':
                output = BytesIO()
                df = pd.DataFrame(data)
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Товары', index=False)
                output.seek(0)
                return output
            
            elif format == 'json':
                output = BytesIO()
                output.write(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
                output.seek(0)
                return output
        
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта: {e}")
            return None
    
    def close(self):
        """Закрыть сессию"""
        self.session.close()